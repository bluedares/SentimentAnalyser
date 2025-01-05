from langchain_openai import ChatOpenAI
import google.generativeai as genai
import os
from typing import Dict, Any, List, Optional
from models.review import Review
import pandas as pd
from tabulate import tabulate
from colorama import Fore
from tqdm import tqdm
import time
import threading
import itertools
import sys
from enum import Enum
from rich.console import Console
from rich.table import Table
import json
import random
from rich import box
from rich.style import Style
from typing import Tuple
from io import StringIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ModelType(Enum):
    OPENAI = "openai"
    GEMINI = "gemini"
    BOTH = "both"

class Emotion(Enum):
    SATISFACTION = "Satisfaction"
    HAPPINESS = "Happiness"
    APPRECIATION = "Appreciation"
    EXCITEMENT = "Excitement"
    NEUTRAL = "Neutral"
    DISAPPOINTMENT = "Disappointment"
    FRUSTRATION = "Frustration"
    ANGER = "Anger"
    CONFUSION = "Confusion"

class LoadingSpinner:
    def __init__(self, message="Processing"):
        self.message = message
        self.spinner_chars = itertools.cycle(['⠋', '⠙', '⠹', '⠸', '⠼', '⠴', '⠦', '⠧', '⠇', '⠏'])
        self.running = False
        self.spinner_thread = None

    def spin(self):
        while self.running:
            sys.stdout.write(f'\r{self.message} {next(self.spinner_chars)}')
            sys.stdout.flush()
            time.sleep(0.1)
        sys.stdout.write('\r' + ' ' * (len(self.message) + 2) + '\r')
        sys.stdout.flush()

    def start(self):
        self.running = True
        self.spinner_thread = threading.Thread(target=self.spin)
        self.spinner_thread.start()

    def stop(self):
        self.running = False
        if self.spinner_thread:
            self.spinner_thread.join()

class SentimentAnalyzer:
    def __init__(self, model_type: ModelType = ModelType.BOTH):
        """Initialize the analyzer with specified model type"""
        self.model_type = model_type
        
        # Initialize models based on selection
        if model_type in [ModelType.OPENAI, ModelType.BOTH]:
            self.openai_model = ChatOpenAI(
                model_name="gpt-3.5-turbo",
                temperature=0
            )
        
        if model_type in [ModelType.GEMINI, ModelType.BOTH]:
            genai.configure(api_key=os.getenv('GOOGLE_API_KEY'))
            self.gemini_model = genai.GenerativeModel('gemini-pro')
        
        # Emotion mappings
        self.emotion_mappings = {
            # Satisfaction group
            'satisfied': Emotion.SATISFACTION,
            'content': Emotion.SATISFACTION,
            'pleased': Emotion.SATISFACTION,
            'satisfaction': Emotion.SATISFACTION,
            
            # Happiness group
            'happy': Emotion.HAPPINESS,
            'joy': Emotion.HAPPINESS,
            'delight': Emotion.HAPPINESS,
            'happiness': Emotion.HAPPINESS,
            'cheerful': Emotion.HAPPINESS,
            
            # Appreciation group
            'grateful': Emotion.APPRECIATION,
            'thankful': Emotion.APPRECIATION,
            'appreciative': Emotion.APPRECIATION,
            'appreciation': Emotion.APPRECIATION,
            
            # Excitement group
            'excited': Emotion.EXCITEMENT,
            'enthusiastic': Emotion.EXCITEMENT,
            'thrilled': Emotion.EXCITEMENT,
            'excitement': Emotion.EXCITEMENT,
            
            # Neutral group
            'neutral': Emotion.NEUTRAL,
            'indifferent': Emotion.NEUTRAL,
            'okay': Emotion.NEUTRAL,
            'moderate': Emotion.NEUTRAL,
            
            # Disappointment group
            'disappointed': Emotion.DISAPPOINTMENT,
            'letdown': Emotion.DISAPPOINTMENT,
            'dissatisfied': Emotion.DISAPPOINTMENT,
            'disappointment': Emotion.DISAPPOINTMENT,
            
            # Frustration group
            'frustrated': Emotion.FRUSTRATION,
            'annoyed': Emotion.FRUSTRATION,
            'irritated': Emotion.FRUSTRATION,
            'frustration': Emotion.FRUSTRATION,
            
            # Anger group
            'angry': Emotion.ANGER,
            'upset': Emotion.ANGER,
            'mad': Emotion.ANGER,
            'anger': Emotion.ANGER,
            
            # Confusion group
            'confused': Emotion.CONFUSION,
            'uncertain': Emotion.CONFUSION,
            'unsure': Emotion.CONFUSION,
            'confusion': Emotion.CONFUSION,
        }
        
        # Define the sentiment analysis prompt template
        emotions_list = ", ".join([e.value for e in Emotion])
        self.prompt_template = """Analyze the emotional content of this review and provide a brief summary.
        Choose EXACTLY ONE emotion from this list: {}
        
        Guidelines:
        1. Select the most appropriate emotion from the provided list
        2. Do not use any emotions outside this list
        3. Provide a brief explanation for the chosen emotion
        
        Review Text: {{}}""".format(emotions_list)
        
        # Rate limiting settings
        self.max_retries = 5
        self.initial_delay = 1  # Initial delay in seconds
        self.max_delay = 32     # Maximum delay in seconds
        self.batch_size = 5     # Number of reviews to process in parallel
        self.delay_between_batches = 2  # Delay between batches in seconds
        
        # Initialize Rich console
        self.console = Console()
        
        # Define styles for different sentiments
        self.sentiment_styles = {
            'Positive': Style(color="green"),
            'Negative': Style(color="red"),
            'Neutral': Style(color="yellow")
        }
        
        # Excel styling
        self.header_style = {
            'font': Font(bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
            'alignment': Alignment(horizontal='center'),
            'border': Border(
                bottom=Side(style='thin'),
                top=Side(style='thin'),
                left=Side(style='thin'),
                right=Side(style='thin')
            )
        }
        
        self.sentiment_colors = {
            'Positive': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
            'Negative': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
            'Neutral': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        }
        
    def _handle_api_error(self, e: Exception, attempt: int, review_author: str) -> Optional[str]:
        """Handle API errors with exponential backoff"""
        if attempt >= self.max_retries:
            print(f"Error analyzing review from {review_author}: Maximum retries exceeded")
            return None
            
        if hasattr(e, 'status_code'):
            if e.status_code == 429:  # Rate limit exceeded
                delay = min(self.initial_delay * (2 ** attempt) + random.uniform(0, 1), self.max_delay)
                print(f"\nRate limit hit, waiting {delay:.2f} seconds before retry...")
                time.sleep(delay)
                return None
            elif e.status_code >= 500:  # Server error
                delay = min(self.initial_delay * (2 ** attempt), self.max_delay)
                time.sleep(delay)
                return None
                
        print(f"Error analyzing review from {review_author}: {str(e)}")
        return None

    def _analyze_with_openai(self, review_text: str, review_author: str, attempt: int = 0) -> Optional[str]:
        """Analyze review using OpenAI with retry logic"""
        try:
            response = self.openai_model.invoke(
                self.prompt_template.format(review_text)
            )
            return response.content
        except Exception as e:
            result = self._handle_api_error(e, attempt, review_author)
            if result is None and attempt < self.max_retries:
                return self._analyze_with_openai(review_text, review_author, attempt + 1)
            return result

    def _analyze_with_gemini(self, review_text: str, review_author: str, attempt: int = 0) -> Optional[str]:
        """Analyze review using Gemini with retry logic"""
        try:
            response = self.gemini_model.generate_content(
                self.prompt_template.format(review_text)
            )
            return response.text
        except Exception as e:
            result = self._handle_api_error(e, attempt, review_author)
            if result is None and attempt < self.max_retries:
                return self._analyze_with_gemini(review_text, review_author, attempt + 1)
            return result

    def analyze_with_openai(self, review: Review) -> Dict[str, Any]:
        """Analyze sentiment using OpenAI's model"""
        response = self._analyze_with_openai(review.text, review.author)
        return {
            "model": "ChatGPT",
            "analysis": response,
            "stars": review.stars,
            "author": review.author,
            "date": review.date,
            "device": review.device,
            "manufacturer": review.manufacturer,
            "version": review.version,
            "text": review.text,
            "client_id": review.client_id
        }

    def analyze_with_gemini(self, review: Review) -> Dict[str, Any]:
        """Analyze sentiment using Google's Gemini model"""
        response = self._analyze_with_gemini(review.text, review.author)
        return {
            "model": "Gemini",
            "analysis": response,
            "stars": review.stars,
            "author": review.author,
            "date": review.date,
            "device": review.device,
            "manufacturer": review.manufacturer,
            "version": review.version,
            "text": review.text,
            "client_id": review.client_id
        }

    def map_to_predefined_emotion(self, text: str) -> Emotion:
        """Map any emotion text to our predefined set of emotions"""
        # Extract the first word (assumed to be the emotion)
        words = text.lower().split()
        if not words:
            return Emotion.NEUTRAL
            
        # Try to find a match in our mappings
        for word in words:
            if word in self.emotion_mappings:
                return self.emotion_mappings[word]
            
            # Try to find partial matches
            for emotion_word, emotion in self.emotion_mappings.items():
                if emotion_word in word or word in emotion_word:
                    return emotion
        
        # Default to neutral if no match found
        return Emotion.NEUTRAL

    def extract_emotion(self, analysis: str) -> str:
        """Extract emotion from analysis text and map to predefined emotion"""
        try:
            # Get the first line or sentence which should contain the emotion
            first_line = analysis.split('\n')[0].split('.')[0]
            mapped_emotion = self.map_to_predefined_emotion(first_line)
            return mapped_emotion.value
        except:
            return Emotion.NEUTRAL.value

    def determine_sentiment(self, rating: float, emotion: str) -> str:
        """Determine sentiment based on emotion first, then rating"""
        # Define emotion categories
        negative_emotions = {'anger', 'frustration', 'disappointment', 'sadness', 'disgust'}
        positive_emotions = {'satisfaction', 'happiness', 'appreciation', 'excitement'}
        
        # First check emotion
        emotion_lower = emotion.lower()
        if emotion_lower in negative_emotions:
            return 'Negative'
        elif emotion_lower in positive_emotions:
            return 'Positive'
        
        # If emotion is neutral (or not in our lists), then use rating
        if rating <= 2:
            return 'Negative'
        elif rating == 3:
            return 'Neutral'
        else:
            return 'Positive'

    def analyze_reviews(self, reviews_text: str) -> pd.DataFrame:
        """Analyze multiple reviews and return a DataFrame with results"""
        # Parse reviews
        print(f"{Fore.CYAN}Parsing reviews...{Fore.WHITE}")
        reviews = []
        current_review = []
        
        for line in reviews_text.split('\n'):
            if line.strip():
                current_review.append(line)
            elif current_review:
                reviews.append(Review.parse_review('\n'.join(current_review)))
                current_review = []
        
        if current_review:
            reviews.append(Review.parse_review('\n'.join(current_review)))

        # Analyze each review
        results = []
        spinner = LoadingSpinner("Analyzing reviews")
        spinner.start()
        
        try:
            for review in reviews:
                try:
                    # Get analysis based on selected model
                    if self.model_type in [ModelType.OPENAI, ModelType.BOTH]:
                        openai_analysis = self.analyze_with_openai(review)
                        results.append(openai_analysis)
                    
                    if self.model_type in [ModelType.GEMINI, ModelType.BOTH]:
                        gemini_analysis = self.analyze_with_gemini(review)
                        results.append(gemini_analysis)
                        
                except Exception as e:
                    print(f"{Fore.RED}Error analyzing review from {review.author}: {str(e)}{Fore.WHITE}")
        finally:
            spinner.stop()

        # Convert to DataFrame
        print(f"{Fore.GREEN}Analysis complete!{Fore.WHITE}")
        df = pd.DataFrame(results)
        return df

    def save_excel_report(self, summary_data, sentiment_data, emotion_data, user_data):
        """Save analysis results to an Excel file with formatting"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Analysis Results"
        
        current_row = 1
        max_col_used = 1
        
        def write_table(ws, title, headers, data, start_row, is_summary=False):
            nonlocal max_col_used
            # Write title with background color
            title_cell = ws.cell(row=start_row, column=1)
            title_cell.value = title
            title_cell.font = Font(bold=True, size=12, color="FFFFFF")  # White text
            title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Dark blue background
            title_cell.alignment = Alignment(horizontal='center')
            
            if is_summary:
                # Merge the two columns for the title
                ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=2)
                max_col_used = max(max_col_used, 2)
                
                # For summary, write data directly under title
                for row_idx, (label, value) in enumerate(data):
                    # Label cell
                    label_cell = ws.cell(row=start_row + row_idx + 1, column=1)
                    label_cell.value = label
                    label_cell.font = Font(bold=True)
                    label_cell.alignment = Alignment(horizontal='center')
                    
                    # Value cell
                    value_cell = ws.cell(row=start_row + row_idx + 1, column=2)
                    value_cell.value = value
                    value_cell.alignment = Alignment(horizontal='center')
                    
                    # Add border and background
                    for cell in [label_cell, value_cell]:
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        cell.fill = PatternFill(start_color="E8F3FF", end_color="E8F3FF", fill_type="solid")
                
                return start_row + len(data) + 1
            else:
                # Update max columns used
                max_col_used = max(max_col_used, len(headers))
                
                # Merge cells for the title
                ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(headers))
                
                # Write headers
                header_row = start_row + 1
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=header_row, column=col)
                    cell.value = header
                    cell.font = self.header_style['font']
                    cell.fill = self.header_style['fill']
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = self.header_style['border']
                
                # Write data
                for row_idx, row_data in enumerate(data, header_row + 1):
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.value = value
                        cell.alignment = Alignment(horizontal='center')
                        
                        # Apply sentiment colors
                        if headers[col_idx-1].lower() == 'sentiment':
                            cell.fill = self.sentiment_colors.get(str(value), PatternFill())
                        
                        # Add border
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                
                return row_idx + 2
        
        # Write Summary first
        current_row = write_table(
            ws=ws,
            title="Summary",
            headers=[],  # No headers for summary
            data=summary_data,
            start_row=current_row,
            is_summary=True
        )
        
        # Add spacing
        current_row += 1
        
        # Write Sentiment Summary
        current_row = write_table(
            ws=ws,
            title="Sentiment Distribution",
            headers=['Sentiment', 'Count', 'Percentage'],
            data=sentiment_data,
            start_row=current_row
        )
        
        # Add spacing
        current_row += 1
        
        # Write Emotion Summary
        if emotion_data:
            current_row = write_table(
                ws=ws,
                title="Emotion Distribution",
                headers=['Emotion', 'Count', 'Sentiment'],
                data=emotion_data,
                start_row=current_row
            )
            
            # Add spacing
            current_row += 1
        
        # Write User Details
        current_row = write_table(
            ws=ws,
            title="User Information",
            headers=['Name', 'Device', 'Client ID', 'Emotion', 'Sentiment', 'Rating', 'Review Preview'],
            data=user_data,
            start_row=current_row
        )
        
        # Hide unused columns and rows
        self.cleanup_worksheet(ws, current_row, max_col_used)
        
        # Save the workbook
        excel_file = 'review_analysis.xlsx'
        wb.save(excel_file)
        return excel_file

    def display_analysis(self, df: pd.DataFrame) -> None:
        """Display the analysis results in formatted tables"""
        # Initialize content for text file
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        file_content = [f"Analysis Time: {timestamp}\n"]
        
        # Process emotions and sentiments
        print(f"\n{Fore.CYAN}Processing emotions...{Fore.WHITE}")
        emotions = []
        emotion_by_review = {}
        sentiment_by_review = {}
        sentiment_counts = {'Positive': 0, 'Negative': 0, 'Neutral': 0}
        
        # First pass: Process emotions and determine sentiments
        for _, row in df.iterrows():
            try:
                emotion = self.extract_emotion(row['analysis'])
                emotions.append(emotion.lower())
                review_key = (row['author'], row['text'])
                emotion_by_review[review_key] = emotion
                sentiment = self.determine_sentiment(row['stars'], emotion)
                sentiment_by_review[review_key] = sentiment
                sentiment_counts[sentiment] += 1
            except Exception as e:
                print(f"Error processing row: {e}")
                continue
        
        # Calculate summary statistics
        total_reviews = len(df)
        avg_rating = df['stars'].mean()
        
        # Add summary at the top for text file
        file_content.append("Summary:")
        file_content.append(f"Number of Reviews: {total_reviews}")
        file_content.append(f"Average Rating: {avg_rating:.2f} stars\n")
        
        # Display sentiment summary
        print(f"\n{Fore.CYAN}=== Sentiment Summary ==={Fore.WHITE}")
        sentiment_data = []
        for sentiment, count in sentiment_counts.items():
            percentage = (count / total_reviews * 100) if total_reviews > 0 else 0
            sentiment_data.append([sentiment, count, f"{percentage:.1f}%"])
        
        # Add sentiment distribution to text file
        file_content.append("\nSentiment Distribution:")
        output = StringIO()
        sentiment_df = pd.DataFrame(sentiment_data, columns=['Sentiment', 'Count', 'Percentage'])
        sentiment_df.to_csv(output, sep='\t', index=False)
        file_content.append(output.getvalue())
        
        # Process emotions
        emotion_data = []
        if emotion_counts := {e: emotions.count(e) for e in set(emotions)}:
            print(f"\n{Fore.CYAN}=== Emotion Summary ==={Fore.WHITE}")
            
            for emotion, count in sorted(emotion_counts.items(), key=lambda x: x[1], reverse=True):
                sample_review = next((key for key, val in emotion_by_review.items() if val.lower() == emotion.lower()), None)
                sentiment = sentiment_by_review.get(sample_review, "Neutral")
                emotion_data.append([emotion.capitalize(), count, sentiment])
            
            # Add emotion distribution to text file
            file_content.append("\nEmotion Distribution:")
            output = StringIO()
            emotion_df = pd.DataFrame(emotion_data, columns=['Emotion', 'Count', 'Sentiment'])
            emotion_df.to_csv(output, sep='\t', index=False)
            file_content.append(output.getvalue())
        
        # Process user details
        print(f"\n{Fore.CYAN}=== User Details ==={Fore.WHITE}")
        user_data = []
        
        for _, row in df.iterrows():
            client_id = row.get('client_id', '')
            if client_id:
                client_id = f" (Client ID: {client_id})"
            
            emotion = self.extract_emotion(row['analysis'])
            review_key = (row['author'], row['text'])
            sentiment = sentiment_by_review.get(review_key, "Neutral")
            
            user_data.append([
                row['author'],
                row['device'],
                client_id,
                emotion,
                sentiment,
                row['stars'],
                row['text'][:100] + ('...' if len(row['text']) > 100 else '')
            ])
        
        # Add user information to text file
        file_content.append("\nUser Information:")
        output = StringIO()
        user_df = pd.DataFrame(user_data, columns=['Name', 'Device', 'Client ID', 'Emotion', 'Sentiment', 'Rating', 'Review Preview'])
        user_df.to_csv(output, sep='\t', index=False)
        file_content.append(output.getvalue())
        
        # Save both formats
        self.save_analysis_results('\n'.join(file_content))
        
        # Prepare summary data
        summary_data = [
            ("Number of Reviews:", f"{total_reviews}"),
            ("Average Rating:", f"{avg_rating:.2f} stars")
        ]
        
        # Save Excel with summary at top
        excel_file = self.save_excel_report(summary_data, sentiment_data, emotion_data, user_data)
        print(f"\n{Fore.GREEN}Analysis saved to review_analysis.txt and {excel_file}{Fore.WHITE}")

    def format_rich_table(self, data: List[List], headers: List[str], title: str = None) -> Table:
        """Format data into a Rich table for console display"""
        table = Table(show_header=True, header_style="bold cyan", title=title)
        
        # Add columns
        for header in headers:
            table.add_column(header, justify="left")
        
        # Add rows
        for row in data:
            styled_row = []
            for i, cell in enumerate(row):
                cell_str = str(cell)
                if headers[i].lower() == 'sentiment':
                    color = {
                        'Positive': 'green',
                        'Negative': 'red',
                        'Neutral': 'yellow'
                    }.get(cell_str, 'white')
                    styled_row.append(f"[{color}]{cell_str}[/{color}]")
                else:
                    styled_row.append(cell_str)
            table.add_row(*styled_row)
        
        return table

    def printTable(self, table, align="LLLL", hasHeader=True):
        """Print formatted table with custom alignment
        align: string of L/R/C chars for Left/Right/Center alignment of each column
        hasHeader: whether first row is a header (will be separated by a line)
        """
        # Find maximum width of each column
        cols = len(table[0])
        widths = []
        for col in range(cols):
            width = max(len(str(row[col])) for row in table)
            widths.append(width)
        
        # Create format string based on alignment
        formats = []
        for i, a in enumerate(align):
            if a == 'L':
                formats.append('{:<' + str(widths[i]) + '}')
            elif a == 'R':
                formats.append('{:>' + str(widths[i]) + '}')
            else:  # Center
                formats.append('{:^' + str(widths[i]) + '}')
        
        # Create separator lines
        separator = '┌' + '┬'.join('─' * w for w in widths) + '┐'
        separator_mid = '├' + '┼'.join('─' * w for w in widths) + '┤'
        separator_bottom = '└' + '┴'.join('─' * w for w in widths) + '┘'
        
        # Print table
        result = []
        result.append(separator)
        
        for i, row in enumerate(table):
            formatted_row = []
            for j, cell in enumerate(row):
                formatted_row.append(formats[j].format(str(cell)))
            result.append('│' + '│'.join(formatted_row) + '│')
            if i == 0 and hasHeader:
                result.append(separator_mid)
        
        result.append(separator_bottom)
        return '\n'.join(result)

    def save_analysis_results(self, content: str, filename: str = "review_analysis.txt", mode: str = 'w') -> None:
        """Save the analysis content to a file"""
        try:
            with open(filename, mode, encoding='utf-8') as f:
                f.write(content)
            if mode == 'w':
                print(f"\n{Fore.GREEN}Analysis results saved to {filename}{Fore.WHITE}")
        except Exception as e:
            print(f"\n{Fore.RED}Error saving to file: {str(e)}{Fore.WHITE}")

    def cleanup_worksheet(self, ws, current_row, max_col_used):
        # Hide unused columns
        for col in range(max_col_used + 1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].hidden = True
        
        # Hide unused rows
        for row in range(current_row + 1, ws.max_row + 1):
            ws.row_dimensions[row].hidden = True
        
        # Adjust column widths for visible columns
        for col in range(1, max_col_used + 1):
            column = get_column_letter(col)
            max_length = 0
            for cell in ws[column]:
                if not ws.row_dimensions[cell.row].hidden:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
